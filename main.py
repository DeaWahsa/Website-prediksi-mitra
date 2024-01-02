import os
import shutil
from typing import Dict, Union
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, declarative_base
import pandas as pd
from database import get_db
from sqlalchemy import table, MetaData, create_engine, select, func, extract, delete
from klasifikasi import function_klasifikasi
from data import Mitra, User
from pydantic import BaseModel
import uvicorn
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
import matplotlib.pyplot as plt
from fastapi.responses import StreamingResponse, FileResponse
from sklearn.preprocessing import LabelEncoder
import logging
import matplotlib
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import json
from pydantic import BaseModel
from datetime import datetime
from typing import List
from typing import Optional



app = FastAPI()

# Konfigurasi CORS
origins = [
    "*"
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/data", StaticFiles(directory="data"), name="static")

logging.basicConfig(level=logging.INFO)

matplotlib.use('Agg')

class UserLogin(BaseModel):
    email: str
    password: str
    
class UserRegistrasi(BaseModel):
    nik: str
    email: str
    password: str
    
# Definisikan Pydantic model yang sesuai dengan struktur data yang Anda ingin serialisasi
class CombinedDataModel(BaseModel):
    tanggal_upload: datetime

    class Config:
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

@app.post("/login")
async def login(user_info: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == user_info.email, User.password == user_info.password).first()
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return {"message": "Login successful", "user_id": user.id}

# Endpoint untuk Register
@app.post("/registrasi", response_model=Dict[str, str])
def register(user: UserRegistrasi, db: Session = Depends(get_db)):
    # Check apakah email sudah terdaftar
    existing_user = db.query(User).filter(User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Email is already registered")

    new_user = User(nik=user.nik, email=user.email, password=user.password)
    db.add(new_user)

    try:
        db.commit()
        db.refresh(new_user)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail="An error occurred while saving to the database. Please try again.")

    # Cek lagi setelah commit untuk memastikan data tersimpan
    check_user = db.query(User).filter(User.email == user.email).first()
    if not check_user:
        raise HTTPException(status_code=500, detail="Registration failed. User data was not saved to the database.")

    return {"message": "User registered successfully!"}

def create_excel_template_GAPED():
    # Membuat workbook baru
    wb = openpyxl.Workbook()
    ws = wb.active

    # Menentukan header
    headers = ['Name of Vendor','Short Text', 'PO Date', 'Delivery Date', 'Doc Date GR', 'Local Amount']

    # Menyisipkan header ke dalam worksheet dan memberi gaya font bold
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws['{}1'.format(col_letter)].font = Font(bold=True)

    # Menyimpan workbook ke dalam io.BytesIO agar bisa langsung dikirimkan sebagai response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_excel_template_SMILE():
    # Membuat workbook baru
    wb = openpyxl.Workbook()
    ws = wb.active

    # Menentukan header
    headers = ['Jumlah Projek', 'Nilai Performansi KHS']

    # Menyisipkan header ke dalam worksheet dan memberi gaya font bold
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws['{}1'.format(col_letter)].font = Font(bold=True)

    # Menyimpan workbook ke dalam io.BytesIO agar bisa langsung dikirimkan sebagai response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_excel_template_Witel():
    # Membuat workbook baru
    wb = openpyxl.Workbook()
    ws = wb.active

    # Menentukan header
    headers = ['Alker/Salker', 'Stok Material','Jumlah Team','Kerapihan']

    # Menyisipkan header ke dalam worksheet dan memberi gaya font bold
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws['{}1'.format(col_letter)].font = Font(bold=True)

    # Menyimpan workbook ke dalam io.BytesIO agar bisa langsung dikirimkan sebagai response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def process_and_save_file(uploaded_file: UploadFile):
    file_location = f"./data/{uploaded_file.filename}"
    with open(file_location, "wb") as file_object:
        shutil.copyfileobj(uploaded_file.file, file_object)
    return file_location

def combine_and_save_files(file1_path, file2_path, file3_path, combined_file_path):
    # Baca data dari ketiga file CSV
    df1 = pd.read_csv(file1_path)
    df2 = pd.read_csv(file2_path)
    df3 = pd.read_csv(file3_path)

    # Gabungkan DataFrames dari ketiga file ke samping (axis=1)
    combined_df = pd.concat([df1, df2, df3], axis=1)

    # Simpan DataFrame hasil gabungan ke dalam file CSV
    combined_df.to_csv(combined_file_path, index=False)

@app.post("/upload")
async def upload_files(background_task: BackgroundTasks, db: Session = Depends(get_db),
                       file1: UploadFile = File(...), file2: UploadFile = File(...), file3: UploadFile = File(...)):
    file1_path = process_and_save_file(file1)
    file2_path = process_and_save_file(file2)
    file3_path = process_and_save_file(file3)
    combined_file_path = f"./data/combined.csv"
    
    # Gabungkan file dan simpan ke dalam file ketiga
    combine_and_save_files(file1_path, file2_path, file3_path, combined_file_path)

    # Baca data dari file gabungan dan masukkan ke dalam tabel Mitra
    combined_df = pd.read_csv(combined_file_path)
    for index, row in combined_df.iterrows():
        db_data = Mitra(
            nof=row.get('Name of Vendor', ''),
            st=row.get('Short Text', ''),
            po=row.get('PO Date', ''),
            dd=row.get('Delivery Date', ''),
            ddgr=row.get('Doc Date GR', ''),
            la=row.get('Local Amount', ''),
            jp=row.get('Jumlah Projek', ''),
            khs=row.get('Nilai Performansi KHS', ''),
            alker=row.get('Alker/Salker', ''),
            stok=row.get('Stok Material', ''),
            tim=row.get('Jumlah Team', ''),
            rapih=row.get('Kerapihan', '')
        )
        db.add(db_data)
    db.commit()

    return {"message": "uploaded successfully"}

def to_dict(obj):
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}

# Endpoint untuk mendapatkan data berdasarkan tahun
@app.get("/get_combined_data")
def get_combined_data(tahun: Optional[int] = Query(None, alias='year'), db: Session = Depends(get_db)):
    if tahun is not None:
        query = select(Mitra).where(extract('year', Mitra.tanggal_upload) == tahun)
    else:
        query = select(Mitra)

    try:
        result = db.execute(query).scalars().all()
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/delete_data_by_upload_year")
def delete_data_by_upload_year(upload_year: int, db: Session = Depends(get_db)):
    try:
        query = delete(Mitra).where(extract('year', Mitra.tanggal_upload) == upload_year)
        result = db.execute(query)
        db.commit()

        if result.rowcount > 0:
            return {"message": f"{result.rowcount} rows deleted successfully"}
        else:
            return {"message": "No data found for the given upload year", "deleted": result.rowcount}
    except Exception as e:
        db.rollback()
        # Log the exception to the console or a file
        print(f"An error occurred: {e}")
        # Provide a more descriptive message in the HTTPException
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred: {str(e)}")

@app.get("/download_filtered_data/{label}")
async def download_filtered_data(label: int):
    # Membaca file CSV
    df = pd.read_csv('./data/data prediksi mitra.csv')
    
    # Filter data berdasarkan label
    filtered_df = df[df['Label Soft Voting'] == label]
    
    # Mengumpulkan jumlah data untuk setiap 'Name of Vendor'
    summary_df = filtered_df.groupby(['Name of Vendor','Short Text','PO Date','Delivery Date','Doc Date GR','Local Amount','Jumlah Projek','Nilai Performansi KHS','Alker/Salker','Stok Material','Jumlah Team','Kerapihan','Durasi Kontrak','Durasi Penyelesaian','nilai projek per LoP','Kategori Anggaran Projek','Kategori Durasi Projek']).size().reset_index(name='Jumlah Projek per Mitra')
    
    # Menyimpan dataframe terfilter ke dalam buffer
    buffer = io.StringIO()
    summary_df.to_csv(buffer, index=False)
    buffer.seek(0)
    
    # Mengirimkan response sebagai file CSV untuk diunduh
    response = StreamingResponse(iter([buffer.getvalue()]),
                                 media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=filtered_data_label_{label}.csv"
    return response

    
@app.get("/get_pie_chart_partner")
async def get_pie_chart_partner(year: int = Query(None, description="Tahun yang dipilih")):
    # Jika tahun tidak diberikan, gunakan tahun saat ini
    current_year = year if year is not None else datetime.now().year
    file_path = f"./data/data_klasifikasi_mitra_{year}.json"
    
    if not os.path.exists(file_path):
        logging.error(f"File {file_path} tidak ditemukan!")
        raise HTTPException(status_code=404, detail=f"File data klasifikasi mitra untuk tahun {current_year} belum dihasilkan")
    
    with open(file_path, "r") as json_file:
        data = json.load(json_file)
        
    return data

@app.get("/get_pie_chart_partner_excellent")
async def get_pie_chart_partner_excellent(year: int = Query(None, description="Tahun yang dipilih")):
    # Jika tahun tidak diberikan, gunakan tahun saat ini
    current_year = year if year is not None else datetime.now().year
    file_path = f"./data/hasil_klasifikasi_mitra_baiksekali_{current_year}.json"
    
    if not os.path.exists(file_path):
        logging.error(f"File {file_path} tidak ditemukan!")
        raise HTTPException(status_code=404, detail=f"File data klasifikasi mitra untuk tahun {current_year} belum dihasilkan")
    
    with open(file_path, "r") as json_file:
        data = json.load(json_file)
        
    return data

@app.get("/get_pie_chart_partner_good")
async def get_pie_chart_partner_good(year: int = Query(None, description="Tahun yang dipilih")):
    # Jika tahun tidak diberikan, gunakan tahun saat ini
    current_year = year if year is not None else datetime.now().year
    file_path = f"./data/hasil_klasifikasi_mitra_baik_{current_year}.json"
    
    if not os.path.exists(file_path):
        logging.error(f"File {file_path} tidak ditemukan!")
        raise HTTPException(status_code=404, detail=f"File data klasifikasi mitra untuk tahun {current_year} belum dihasilkan")
    
    with open(file_path, "r") as json_file:
        data = json.load(json_file)
        
    return data

@app.get("/get_pie_chart_partner_fair")
async def get_pie_chart_partner_fair(year: int = Query(None, description="Tahun yang dipilih")):
    # Jika tahun tidak diberikan, gunakan tahun saat ini
    current_year = year if year is not None else datetime.now().year
    file_path = f"./data/hasil_klasifikasi_mitra_cukup_{current_year}.json"
    
    if not os.path.exists(file_path):
        logging.error(f"File {file_path} tidak ditemukan!")
        raise HTTPException(status_code=404, detail=f"File data klasifikasi mitra untuk tahun {current_year} belum dihasilkan")
    
    with open(file_path, "r") as json_file:
        data = json.load(json_file)
        
    return data

@app.get("/get_pie_chart_partner_bad")
async def get_pie_chart_partner_bad(year: int = Query(None, description="Tahun yang dipilih")):
    # Jika tahun tidak diberikan, gunakan tahun saat ini
    current_year = year if year is not None else datetime.now().year
    file_path = f"./data/hasil_klasifikasi_mitra_buruk_{current_year}.json"
    
    if not os.path.exists(file_path):
        logging.error(f"File {file_path} tidak ditemukan!")
        raise HTTPException(status_code=404, detail=f"File data klasifikasi mitra untuk tahun {current_year} belum dihasilkan")
    
    with open(file_path, "r") as json_file:
        data = json.load(json_file)
        
    return data

def baca_data_tahunan(tahun_awal, tahun_akhir):
    data_tahunan = {}
    for tahun in range(tahun_awal, tahun_akhir + 1):
        path_file = f'd:\\xampp\\htdocs\\skripsiku\\api\\data\\data_klasifikasi_mitra_{tahun}.json'
        if os.path.exists(path_file):
            with open(path_file, 'r') as file:
                data = json.load(file)
                data_tahunan[tahun] = data
    return data_tahunan


def olah_data(data_tahunan):
    data_olah = {}
    for tahun, data in data_tahunan.items():
        totals = data['totals']
        data_olah[tahun] = totals
    return data_olah

def simpan_data_json(data, path_file):
    with open(path_file, 'w') as file:
        json.dump(data, file, indent=4)
        
@app.get("/get_data_tahunan")
async def get_data_tahunan(tahun: int = Query(..., description="Tahun yang dipilih")):
    path_file = f'd:\\xampp\\htdocs\\skripsiku\\api\\data\\data_klasifikasi_mitra_{tahun}.json'
    if os.path.exists(path_file):
        with open(path_file, 'r') as file:
            data = json.load(file)
        return data
    else:
        raise HTTPException(status_code=404, detail=f"File data untuk tahun {tahun} tidak ditemukan")


def generate_data_klasifikasi_mitra_komulatif():
    tahun_awal = 2020
    tahun_akhir = datetime.now().year

    # Membaca data tahunan
    data_tahunan = baca_data_tahunan(tahun_awal, tahun_akhir)
    data_olah = olah_data(data_tahunan)

    # Lokasi file data kumulatif
    path_simpan = 'd:\\xampp\\htdocs\\skripsiku\\api\\data\\data_klasifikasi_mitra_komulatif.json'

    # Baca data kumulatif yang sudah ada dan gabungkan
    if os.path.exists(path_simpan):
        with open(path_simpan, 'r') as file:
            data_kumulatif = json.load(file)
            for tahun, data in data_olah.items():
                if tahun in data_kumulatif:
                    # Update data tahunan jika sudah ada
                    for key, value in data.items():
                        data_kumulatif[tahun][key] += value
                else:
                    # Tambahkan data tahunan baru
                    data_kumulatif[tahun] = data
    else:
        data_kumulatif = data_olah

    # Simpan data gabungan
    simpan_data_json(data_kumulatif, path_simpan)


@app.get("/data_klasifikasi_mitra_komulatif")
async def read_data_klasifikasi_mitra_komulatif():
    file_path = 'd:\\xampp\\htdocs\\skripsiku\\api\\data\\data_klasifikasi_mitra_komulatif.json'

    # Periksa apakah file ada
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")

    # Baca file JSON dan kirimkan isinya
    with open(file_path, "r") as file:
        data = json.load(file)
        return data


@app.get("/get_data")
def read_root(db: Session = Depends(get_db)):
    return db.query(Mitra).all()

@app.post("/run_klasifikasi")
async def do_klasifikasi(background_task: BackgroundTasks, db: Session = Depends(get_db)):
    data = db.query(Mitra).all()
    try:
        background_task.add_task(function_klasifikasi, data)
        background_task.add_task(generate_data_klasifikasi_mitra_komulatif)

        return {
            'message': 'Running klasifikasi...',
            'data': data
        }
    except Exception as e:
        print(f"Error saat menjalankan klasifikasi: {e}")
        return {
            'message': f"Terjadi kesalahan: {e}",
            'data': []
        }
        
@app.get("/download_excel_template_GAPED/")
async def download_excel_template_GAPED():
    template_io = create_excel_template_GAPED()
    return StreamingResponse(template_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template.xlsx"})
    
@app.get("/download_excel_template_SMILE/")
async def download_excel_template_SMILE():
    template_io = create_excel_template_SMILE()
    return StreamingResponse(template_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template.xlsx"})
    
@app.get("/download_excel_template_Witel/")
async def download_excel_template_Witel():
    template_io = create_excel_template_Witel()
    return StreamingResponse(template_io, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template.xlsx"})

if __name__ == "__main__":
    uvicorn.run("main:app", host='127.0.0.1', port=8081, reload=True)
